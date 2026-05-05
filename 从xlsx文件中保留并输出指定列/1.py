# -*- coding: utf-8 -*-
import pandas as pd
import os
import sys
import configparser
import re
from openpyxl import load_workbook
from openpyxl.styles import numbers
import numpy as np

# ===== 配置文件处理 =====
def load_config(config_path):
    """加载并解析配置文件"""
    config = configparser.ConfigParser()
    try:
        config.read(config_path, encoding='utf-8')
    except:
        try:
            config.read(config_path, encoding='gbk')
        except Exception as e:
            print(f"无法读取配置文件: {str(e)}")
            return None
    
    config_dict = {}
    
    # 解析每个数据集配置
    for section in config.sections():
        # 获取基础配置
        dataset_name = section
        keep_columns = [col.strip() for col in config[section]['keep_columns'].split(',')]
        double_header = config[section].getboolean('double_header', False)
        merge_same_base = config[section].getboolean('merge_same_base', True)
        
        # 获取文件名模式
        file_pattern = config[section].get('file_pattern', '')
        if not file_pattern:
            # 如果没有指定模式，则使用数据集名称作为文件名
            file_pattern = f"{dataset_name}.*"
        
        # 获取文本格式列
        text_columns = []
        if 'text_columns' in config[section]:
            text_columns = [col.strip() for col in config[section]['text_columns'].split(',')]
        
        # 获取长数字ID列
        long_id_columns = []
        if 'long_id_columns' in config[section]:
            long_id_columns = [col.strip() for col in config[section]['long_id_columns'].split(',')]
        
        config_dict[dataset_name] = {
            'keep_columns': keep_columns,
            'text_columns': text_columns,
            'long_id_columns': long_id_columns,
            'double_header': double_header,
            'merge_same_base': merge_same_base,
            'file_pattern': file_pattern
        }
    
    return config_dict

# ===== 文件处理函数 =====
def process_file(file_path, config):
    """处理单个文件"""
    try:
        file_name = os.path.basename(file_path)
        print(f"处理文件: {file_name}")
        
        # 使用openpyxl直接读取Excel文件
        wb = load_workbook(file_path, data_only=True)
        ws = wb.active
        
        # 确定数据起始行
        start_row = 2 if config['double_header'] else 1
        
        # 获取列名映射
        col_names = []
        for col_idx in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=start_row, column=col_idx).value
            col_names.append(str(cell_value) if cell_value is not None else f"Column_{col_idx}")
        
        # 创建列名到索引的映射
        col_index_map = {name: idx for idx, name in enumerate(col_names, 1)}
        
        # 收集数据
        data = []
        for row_idx in range(start_row + 1, ws.max_row + 1):
            row_data = {}
            for col_name in config['keep_columns']:
                if col_name in col_index_map:
                    col_idx = col_index_map[col_name]
                    cell = ws.cell(row=row_idx, column=col_idx)
                    
                    # 特殊处理长数字ID列
                    if col_name in config['long_id_columns']:
                        # 尝试从原始XML中获取原始值
                        if cell.data_type == 'n' and cell.value is not None:
                            # 对于数值型长ID，使用格式化字符串保持精度
                            value = f"{cell.value:.0f}" if cell.value.is_integer() else str(cell.value)
                        else:
                            value = str(cell.value) if cell.value is not None else None
                    else:
                        value = cell.value
                    
                    row_data[col_name] = value
            data.append(row_data)
        
        # 创建DataFrame
        df = pd.DataFrame(data)
        
        # 添加文件名标识列
        if not df.empty:
            df['_源文件名'] = file_name
        
        return df
        
    except Exception as e:
        print(f"处理 {file_name} 失败: {str(e)}")
        return None

# ===== 保存函数 =====
def save_with_text_format(df, output_path, text_columns, long_id_columns):
    """保存Excel并确保指定列保持文本格式"""
    try:
        # 先保存为临时Excel
        temp_path = output_path.replace(".xlsx", "_temp.xlsx")
        df.to_excel(temp_path, index=False)
        
        # 使用openpyxl加载工作簿
        wb = load_workbook(temp_path)
        ws = wb.active
        
        # 获取列索引映射
        col_index_map = {}
        for idx, col in enumerate(df.columns, 1):
            col_index_map[col] = idx
        
        # 设置文本格式
        for col_name in text_columns + long_id_columns:
            if col_name in col_index_map:
                col_idx = col_index_map[col_name]
                # 设置整列为文本格式
                for row in range(2, len(df) + 2):  # 从第2行开始（跳过标题行）
                    cell = ws.cell(row=row, column=col_idx)
                    # 确保值作为字符串保存
                    if cell.value is not None:
                        cell.value = str(cell.value)
                        cell.number_format = numbers.FORMAT_TEXT
        
        # 保存最终文件
        wb.save(output_path)
        wb.close()
        
        # 删除临时文件
        os.remove(temp_path)
        
        return True
        
    except Exception as e:
        print(f"保存文件失败: {output_path} | 错误: {str(e)}")
        # 尝试直接保存
        df.to_excel(output_path, index=False)
        return False

# ===== 主程序 =====
def main():
    # 设置路径
    current_dir = os.path.dirname(os.path.abspath(__file__))
    config_path = os.path.join(current_dir, "config.txt")
    input_dir = os.path.join(current_dir, "输入")
    output_dir = os.path.join(current_dir, "输出")
    
    # 确保目录存在
    os.makedirs(input_dir, exist_ok=True)
    os.makedirs(output_dir, exist_ok=True)
    
    print("=" * 50)
    print("开始数据清洗处理...")
    print(f"配置文件: {config_path}")
    print(f"输入目录: {input_dir}")
    print(f"输出目录: {output_dir}")
    print("=" * 50)
    
    # 加载配置
    config_dict = load_config(config_path)
    if not config_dict:
        print("错误: 无法加载配置文件，请检查config.txt格式")
        return
    
    print(f"加载配置: 共 {len(config_dict)} 个数据集配置")
    
    # 收集所有输入文件
    all_files = []
    for file in os.listdir(input_dir):
        if file.lower().endswith('.xlsx'):
            all_files.append(os.path.join(input_dir, file))
    
    if not all_files:
        print("警告: 输入目录中没有找到Excel文件")
        return
    
    # 按配置处理文件
    processed_files = 0
    merged_datasets = {}
    
    for dataset_name, config in config_dict.items():
        print(f"\n处理数据集: {dataset_name}")
        print(f"  保留列: {', '.join(config['keep_columns'])}")
        if config['text_columns']:
            print(f"  文本格式列: {', '.join(config['text_columns'])}")
        if config['long_id_columns']:
            print(f"  长数字ID列: {', '.join(config['long_id_columns'])}")
        print(f"  双行表头: {'是' if config['double_header'] else '否'}")
        print(f"  合并相同数据集: {'是' if config['merge_same_base'] else '否'}")
        
        # 匹配符合模式的文件
        pattern = re.compile(config['file_pattern'].replace('.', r'\.').replace('*', '.*'))
        matched_files = [f for f in all_files if pattern.search(os.path.basename(f))]
        
        if not matched_files:
            print(f"  警告: 没有找到匹配模式 '{config['file_pattern']}' 的文件")
            continue
        
        # 处理匹配的文件
        dataset_dfs = []
        for file_path in matched_files:
            df = process_file(file_path, config)
            if df is not None and not df.empty:
                dataset_dfs.append(df)
                processed_files += 1
        
        if not dataset_dfs:
            print(f"  警告: 数据集 '{dataset_name}' 没有成功处理的文件")
            continue
        
        # 合并相同数据集
        if config['merge_same_base'] and len(dataset_dfs) > 1:
            merged_df = pd.concat(dataset_dfs, ignore_index=True)
            output_file = f"cleaned_{dataset_name}.xlsx"
            output_path = os.path.join(output_dir, output_file)
            
            # 保存并设置文本格式
            save_with_text_format(merged_df, output_path, config['text_columns'], config['long_id_columns'])
            
            print(f"  合并完成: {len(dataset_dfs)}个文件 → {len(merged_df)}行 → 保存到 {output_file}")
            merged_datasets[dataset_name] = output_path
        else:
            # 单独保存每个文件
            for i, df in enumerate(dataset_dfs):
                source_file = os.path.basename(df['_源文件名'].iloc[0])
                base_name = os.path.splitext(source_file)[0]
                
                # 对于命名规则文件，使用数据集名称
                if config['merge_same_base'] and '_' in base_name:
                    base_part = base_name.split('_')[0]
                    if base_part == dataset_name:
                        output_file = f"cleaned_{dataset_name}.xlsx"
                    else:
                        output_file = f"cleaned_{base_name}.xlsx"
                else:
                    output_file = f"cleaned_{base_name}.xlsx"
                
                output_path = os.path.join(output_dir, output_file)
                
                # 保存并设置文本格式
                save_with_text_format(df, output_path, config['text_columns'], config['long_id_columns'])
                
                print(f"  保存: {source_file} → {output_file} ({len(df)}行)")
    
    print("\n" + "=" * 50)
    print(f"处理完成! 成功处理 {processed_files} 个文件")
    
    if merged_datasets:
        print("\n合并的数据集:")
        for name, path in merged_datasets.items():
            print(f"  {name}: {os.path.basename(path)}")
    
    print("=" * 50)

# 程序入口
if __name__ == "__main__":
    main()