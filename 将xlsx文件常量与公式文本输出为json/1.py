#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
提取 Excel/CSV 文件中所有非空单元格的内容，输出增强结构的 JSON 文件。
每个单元格包含类型（constant/formula/error/date）和值（公式已清理前缀）。
溢出区域自动忽略。
依赖库：openpyxl, xlrd (推荐版本 1.2.0), 标准库 csv, json, os, sys, re
"""

import os
import sys
import json
import csv
import re
from datetime import datetime, date
from collections import OrderedDict

# 尝试导入所需库，若缺失则提示安装
try:
    import openpyxl
    from openpyxl.worksheet.formula import ArrayFormula
except ImportError:
    print("缺少 openpyxl 库，请运行：pip install openpyxl")
    sys.exit(1)

try:
    import xlrd
except ImportError:
    print("缺少 xlrd 库，请运行：pip install xlrd==1.2.0")
    sys.exit(1)

# -------------------- 辅助函数 --------------------
def col_letter_to_num(col_letter):
    """将 Excel 列字母转换为数字（A=1, B=2, ..., Z=26, AA=27, ...）"""
    num = 0
    for ch in col_letter:
        num = num * 26 + (ord(ch.upper()) - ord('A') + 1)
    return num

def cell_coord_to_col_row(coord):
    """将 Excel 坐标如 "A2" 转换为 (列号, 行号)，列号从1开始，行号从1开始"""
    col_letter = ''.join([c for c in coord if not c.isdigit()])
    row_num = int(''.join([c for c in coord if c.isdigit()]))
    col_num = col_letter_to_num(col_letter)
    return col_num, row_num

def parse_ref(ref):
    """解析类似 "A2:H1643" 的引用，返回起始列字母、起始行、结束列字母、结束行"""
    parts = ref.split(':')
    start = parts[0]
    end = parts[1]
    start_col = ''.join([c for c in start if not c.isdigit()])
    start_row = int(''.join([c for c in start if c.isdigit()]))
    end_col = ''.join([c for c in end if not c.isdigit()])
    end_row = int(''.join([c for c in end if c.isdigit()]))
    return start_col, start_row, end_col, end_row

def col_letter(n):
    """将列索引（从1开始）转换为Excel列字母（A, B, ..., Z, AA, ...）"""
    letters = ""
    while n > 0:
        n -= 1
        letters = chr(65 + (n % 26)) + letters
        n //= 26
    return letters

def cell_coord(row, col):
    """根据行号（从1开始）和列号（从1开始）返回Excel坐标字符串，如 "A1" """
    return f"{col_letter(col)}{row}"

def safe_json_serialize(obj):
    """将无法直接JSON序列化的对象（如日期）转换为字符串"""
    if isinstance(obj, (datetime, date)):
        return obj.isoformat()
    return str(obj)

def clean_formula(formula):
    """
    去除公式中的 _xlfn. 和 _xlpm. 前缀。
    例如：'=_xlfn.SEQUENCE(5,1,2025,1)' -> '=SEQUENCE(5,1,2025,1)'
          '=_xlpm.提取数据' -> '=提取数据'
    """
    if not isinstance(formula, str):
        return formula
    # 确保公式以等号开头（但保留等号）
    cleaned = re.sub(r'_xlfn\.|_xlpm\.|_xlws.', '', formula)
    return cleaned

# -------------------- 处理器类 --------------------
class ExcelProcessor:
    """处理 .xlsx 文件（使用 openpyxl）- 增强版，支持忽略溢出区域并输出新结构"""
    @staticmethod
    def process(filepath):
        wb = openpyxl.load_workbook(filepath, data_only=False)
        workbook_name = os.path.basename(filepath)
        sheets_data = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            # --- 第一步：收集所有动态数组公式的区域信息 ---
            array_ranges = []  # 每个元素为字典：{min_row, max_row, min_col, max_col, tl_row, tl_col}
            for row in ws.iter_rows():
                for cell in row:
                    if cell.data_type == 'f' and isinstance(cell.value, ArrayFormula):
                        ref = getattr(cell.value, 'ref', None)
                        if ref:
                            try:
                                start_col, start_row, end_col, end_row = parse_ref(ref)
                                min_col = col_letter_to_num(start_col)
                                max_col = col_letter_to_num(end_col)
                                tl_col, tl_row = cell_coord_to_col_row(cell.coordinate)
                                array_ranges.append({
                                    'min_row': start_row,
                                    'max_row': end_row,
                                    'min_col': min_col,
                                    'max_col': max_col,
                                    'tl_row': tl_row,
                                    'tl_col': tl_col
                                })
                            except Exception:
                                # 如果解析失败，忽略该区域（安全降级）
                                pass

            # --- 第二步：遍历单元格，收集数据 ---
            rows_dict = {}  # key: 行号, value: 该行单元格列表
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is None:
                        continue
                    coord = cell.coordinate
                    cur_col, cur_row = cell_coord_to_col_row(coord)

                    # 判断是否为溢出区域
                    is_spill = False
                    for r in array_ranges:
                        if (r['min_row'] <= cur_row <= r['max_row'] and
                            r['min_col'] <= cur_col <= r['max_col']):
                            if not (cur_row == r['tl_row'] and cur_col == r['tl_col']):
                                is_spill = True
                                break
                    if is_spill:
                        continue

                    col_letter_str = cell.column_letter
                    row_num = cell.row
                    cell_type = None
                    value = None

                    # 判断单元格类型
                    if cell.data_type == 'f':
                        # 公式单元格
                        cell_type = 'formula'
                        if isinstance(cell.value, ArrayFormula):
                            raw_formula = getattr(cell.value, 'text', str(cell.value))
                        else:
                            raw_formula = cell.value
                        # 确保公式以等号开头（通常已有）
                        if not raw_formula.startswith('='):
                            raw_formula = '=' + raw_formula
                        value = clean_formula(raw_formula)
                    elif cell.data_type == 'e':
                        # 错误值
                        cell_type = 'error'
                        value = cell.value
                    else:
                        # 常量（可能为数字、字符串、日期、布尔）
                        if cell.data_type == 'n' and cell.is_date:
                            cell_type = 'date'
                            value = cell.value  # datetime 对象，后续序列化
                        else:
                            # 统一为 constant，可根据需要细分
                            cell_type = 'constant'
                            value = cell.value

                    # 存入行字典
                    if row_num not in rows_dict:
                        rows_dict[row_num] = []
                    rows_dict[row_num].append({
                        'col': col_letter_str,
                        'type': cell_type,
                        'value': value
                    })

            # 将行字典转换为有序列表
            rows_list = []
            for row_num in sorted(rows_dict.keys()):
                # 按列字母排序单元格
                cells_sorted = sorted(rows_dict[row_num], key=lambda x: x['col'])
                rows_list.append({
                    'row': row_num,
                    'cells': cells_sorted
                })

            sheets_data.append({
                'name': sheet_name,
                'rows': rows_list
            })

        # 可选：收集命名区域（暂不实现，留作扩展）
        named_ranges = []

        result = {
            'workbookName': workbook_name,
            'sheets': sheets_data
        }
        if named_ranges:
            result['namedRanges'] = named_ranges

        wb.close()
        return result

class XlsProcessor:
    """处理 .xls 文件（使用 xlrd 1.2.0）"""
    @staticmethod
    def process(filepath):
        wb = xlrd.open_workbook(filepath, formatting_info=True)
        workbook_name = os.path.basename(filepath)
        sheets_data = []

        for sheet_idx in range(wb.nsheets):
            sheet = wb.sheet_by_index(sheet_idx)
            sheet_name = sheet.name
            rows_dict = {}

            for row in range(sheet.nrows):
                for col in range(sheet.ncols):
                    cell = sheet.cell(row, col)
                    if cell.ctype == xlrd.XL_CELL_EMPTY:
                        continue

                    coord = cell_coord(row + 1, col + 1)
                    col_letter_str = col_letter(col + 1)
                    row_num = row + 1
                    cell_type = None
                    value = None

                    if cell.ctype == xlrd.XL_CELL_FORMULA:
                        cell_type = 'formula'
                        # 获取公式文本
                        try:
                            formula = sheet.cell_formula(row, col)
                        except AttributeError:
                            formula = getattr(cell, 'rdbms_formula', None)
                        if formula:
                            if not formula.startswith('='):
                                formula = '=' + formula
                            value = clean_formula(formula)
                        else:
                            # 降级：使用计算值（但尽量保留公式文本）
                            value = str(cell.value)  # 或标记为错误？
                    elif cell.ctype == xlrd.XL_CELL_DATE:
                        cell_type = 'date'
                        try:
                            date_tuple = xlrd.xldate.xldate_as_tuple(cell.value, wb.datemode)
                            value = date(*date_tuple[:3]).isoformat()
                        except:
                            value = str(cell.value)
                    elif cell.ctype == xlrd.XL_CELL_ERROR:
                        cell_type = 'error'
                        value = xlrd.error_text_from_code.get(cell.value, f"Error {cell.value}")
                    else:
                        cell_type = 'constant'
                        value = cell.value

                    if row_num not in rows_dict:
                        rows_dict[row_num] = []
                    rows_dict[row_num].append({
                        'col': col_letter_str,
                        'type': cell_type,
                        'value': value
                    })

            # 整理行数据
            rows_list = []
            for row_num in sorted(rows_dict.keys()):
                cells_sorted = sorted(rows_dict[row_num], key=lambda x: x['col'])
                rows_list.append({
                    'row': row_num,
                    'cells': cells_sorted
                })
            sheets_data.append({
                'name': sheet_name,
                'rows': rows_list
            })

        result = {
            'workbookName': workbook_name,
            'sheets': sheets_data
        }
        return result

class CsvProcessor:
    """处理 .csv 文件（所有数据视为一个工作表，无公式）"""
    @staticmethod
    def process(filepath, encoding='utf-8'):
        # 尝试自动检测编码（若 utf-8 失败则尝试 gbk）
        try:
            with open(filepath, 'r', encoding=encoding) as f:
                reader = csv.reader(f)
                rows = list(reader)
        except UnicodeDecodeError:
            if encoding == 'utf-8':
                return CsvProcessor.process(filepath, encoding='gbk')
            else:
                raise

        workbook_name = os.path.basename(filepath)
        sheet_name = os.path.splitext(workbook_name)[0]
        rows_dict = {}

        for r, row in enumerate(rows, start=1):
            for c, val in enumerate(row, start=1):
                if val == '':
                    continue
                col_letter_str = col_letter(c)
                cell_type = 'constant'
                value = val
                if r not in rows_dict:
                    rows_dict[r] = []
                rows_dict[r].append({
                    'col': col_letter_str,
                    'type': cell_type,
                    'value': value
                })

        rows_list = []
        for row_num in sorted(rows_dict.keys()):
            cells_sorted = sorted(rows_dict[row_num], key=lambda x: x['col'])
            rows_list.append({
                'row': row_num,
                'cells': cells_sorted
            })
        sheets_data = [{
            'name': sheet_name,
            'rows': rows_list
        }]

        result = {
            'workbookName': workbook_name,
            'sheets': sheets_data
        }
        return result

# -------------------- 主程序 --------------------
def main():
    # 获取脚本所在目录
    script_dir = os.path.dirname(os.path.abspath(__file__))
    input_dir = os.path.join(script_dir, "输入")
    output_dir = os.path.join(script_dir, "输出")

    # 检查输入文件夹是否存在
    if not os.path.isdir(input_dir):
        print(f"错误：输入文件夹不存在 - {input_dir}")
        sys.exit(1)

    # 创建输出文件夹（如果不存在）
    os.makedirs(output_dir, exist_ok=True)

    # 支持的文件扩展名（不区分大小写）
    extensions = ('.xls', '.xlsx', '.csv')

    # 遍历输入文件夹
    for filename in os.listdir(input_dir):
        if not filename.lower().endswith(extensions):
            continue

        filepath = os.path.join(input_dir, filename)
        base_name = os.path.splitext(filename)[0]
        output_file = os.path.join(output_dir, base_name + '.json')

        print(f"正在处理：{filename}")

        try:
            # 根据扩展名选择处理器
            if filename.lower().endswith('.xlsx'):
                result = ExcelProcessor.process(filepath)
            elif filename.lower().endswith('.xls'):
                result = XlsProcessor.process(filepath)
            elif filename.lower().endswith('.csv'):
                result = CsvProcessor.process(filepath)
            else:
                continue

            # 将结果写入 JSON 文件
            with open(output_file, 'w', encoding='utf-8') as f:
                json.dump(result, f, ensure_ascii=False, indent=None, default=safe_json_serialize)

            print(f"已生成：{output_file}")

        except Exception as e:
            print(f"处理文件 {filename} 时出错：{e}")

if __name__ == "__main__":
    main()
